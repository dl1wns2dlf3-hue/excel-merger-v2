import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import re
import io
from datetime import datetime

# --- [핵심 추가] 메모리 폭발 방지 전처리 함수 ---
def safe_read_excel(file_obj):
    """
    무거운 엑셀 파일을 통째로 읽지 않고, read_only 모드로 가볍게 한 줄씩 읽습니다.
    빈 줄이 연속으로 50번 이상 나오면 '유령 데이터'로 간주하고 과감히 잘라냅니다.
    """
    # 메모리 절약을 위한 read_only=True 적용
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    sheets_data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []
        empty_row_count = 0
        
        # 한 줄씩 값을 읽어옴
        for row in ws.iter_rows(values_only=True):
            # 현재 줄이 완전히 비어있는지 확인
            is_empty = all(cell is None or str(cell).strip() == '' for cell in row)
            
            if is_empty:
                empty_row_count += 1
                # 빈 줄이 50번 연속으로 나오면 '끝'으로 간주하고 즉시 읽기 종료! (메모리 세이브)
                if empty_row_count > 50:
                    break
            else:
                empty_row_count = 0
            
            data.append(row)
            
        if data:
            sheets_data[sheet_name] = pd.DataFrame(data) # 판다스 데이터프레임으로 안전하게 변환
            
    wb.close()
    return sheets_data


# --- [공통 함수] 시드 파일에서 데이터 추출 ---
def load_seed_data(file_obj):
    empty_df = pd.DataFrame(columns=['이름_key', '생년월일_val', '업체_val', '직종_val'])
    if file_obj is None: 
        return empty_df
    
    try:
        # 기존에 메모리를 터뜨리던 pd.read_excel 대신, 직접 만든 전처리 함수 사용!
        sheets = safe_read_excel(file_obj)
    except Exception as e:
        st.error(f"파일 전처리 중 오류 발생 ({file_obj.name}): {e}")
        return empty_df
        
    df_list = []
    for sheet_name, df_raw in sheets.items():
        if df_raw is None or df_raw.empty: continue
            
        header_idx = -1
        for idx, row in df_raw.head(20).iterrows():
            row_clean = [str(cell).replace(' ', '').replace('\n', '').strip() for cell in row]
            if '성명' in row_clean or '이름' in row_clean:
                header_idx = idx
                break
        
        if header_idx != -1:
            df = df_raw.iloc[header_idx + 1:].copy()
            cols = [str(c).replace(' ', '').replace('\n', '').strip() for c in df_raw.iloc[header_idx].values]
            df.columns = cols
            df = df.loc[:, ~df.columns.duplicated()].copy()
            df_list.append(df)
            
    if not df_list: 
        return empty_df
    
    combined_df = pd.concat(df_list, ignore_index=True)
    
    name_col = next((c for c in ['이름', '성명'] if c in combined_df.columns), None)
    if not name_col: 
        return empty_df
    
    result_df = pd.DataFrame()
    result_df['이름_key'] = combined_df[name_col].astype(str).str.replace(r'\s+', '', regex=True).str.strip()
    
    def clean_date(x):
        if pd.isna(x) or str(x).strip() in ['', 'nan', 'NaT', 'None']: return ''
        s = str(x).strip()
        if s.endswith('.0'): s = s[:-2]
        s = s.split(' ')[0]
        return re.sub(r'\D', '', s)

    if '생년월일' in combined_df.columns:
        result_df['생년월일_val'] = combined_df['생년월일'].apply(clean_date)
    else:
        result_df['생년월일_val'] = ''
        
    comp_col = next((c for c in ['업체명', '업체', '소속'] if c in combined_df.columns), None)
    result_df['업체_val'] = combined_df[comp_col].astype(str).str.strip().replace('nan', '') if comp_col else ''
    
    job_col = next((c for c in ['직종명', '직종', '공종', '직책'] if c in combined_df.columns), None)
    result_df['직종_val'] = combined_df[job_col].astype(str).str.strip().replace('nan', '') if job_col else ''
    
    result_df = result_df[result_df['이름_key'].notna() & (result_df['이름_key'] != '') & (result_df['이름_key'] != 'nan')]
    return result_df

# --- [웹 UI 설정] ---
st.set_page_config(page_title="근로자 데이터 병합기", page_icon="👷", layout="centered")

st.title("👷 근로자 데이터 자동 병합 시스템")
st.info("시드 파일(1번, 2번) 중 하나만 올려도 작동합니다. 타겟 파일의 빈칸을 자동으로 채워줍니다.")

seed1_file = st.file_uploader("1. 첫 번째 시드 파일 (선택)", type=["xlsx"])
seed2_file = st.file_uploader("2. 두 번째 시드 파일 (선택)", type=["xlsx"])
target_file = st.file_uploader("3. 작성하려는 타겟 파일 (필수)", type=["xlsx"])

if st.button("데이터 병합 실행 🚀"):
    if not target_file:
        st.error("작성하려는 타겟 파일을 업로드해주세요.")
    elif not seed1_file and not seed2_file:
        st.error("데이터를 가져올 시드 파일(1번 혹은 2번)을 최소 하나는 올려주세요.")
    else:
        with st.spinner('무거운 엑셀 파일을 최적화하여 읽어오는 중입니다...'):
            try:
                df_s1 = load_seed_data(seed1_file)
                df_s2 = load_seed_data(seed2_file)
                df_seeds = pd.concat([df_s1, df_s2], ignore_index=True)
                
                if df_seeds.empty:
                    st.error("시드 파일에서 유효한 명단 데이터를 찾지 못했습니다.")
                    st.stop()
                
                df_seeds['score'] = (df_seeds['생년월일_val'] != '').astype(int) + \
                                   (df_seeds['업체_val'] != '').astype(int) + \
                                   (df_seeds['직종_val'] != '').astype(int)
                df_seeds = df_seeds.sort_values('score', ascending=False).drop_duplicates(subset=['이름_key'], keep='first')

                # 타겟 파일 처리
                wb = openpyxl.load_workbook(target_file)
                target_sheet = None
                header_row_idx = 1
                col_indices = {}
                
                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                        row_strs = [str(cell).replace(' ', '').replace('\n', '').strip() if cell else '' for cell in row]
                        if '이름' in row_strs or '성명' in row_strs:
                            target_sheet = ws
                            header_row_idx = row_idx
                            for col_idx, val in enumerate(row_strs, 1):
                                if val: col_indices[val] = col_idx
                            break
                    if target_sheet: break
                        
                if not target_sheet:
                    st.error("타겟 파일에서 '이름' 또는 '성명' 열이 있는 시트를 찾지 못했습니다.")
                    st.stop()

                name_key = '이름' if '이름' in col_indices else '성명'
                t_dob = '생년월일' if '생년월일' in col_indices else None
                t_comp = next((k for k in ['업체명', '업체', '소속'] if k in col_indices), None)
                t_job = next((k for k in ['직종명', '직종', '공종', '직책'] if k in col_indices), None)

                fill_count = 0
                for r in range(header_row_idx + 1, target_sheet.max_row + 1):
                    name_val = target_sheet.cell(row=r, column=col_indices[name_key]).value
                    if not name_val: continue
                    
                    name = str(name_val).replace(' ', '').replace('\n', '').strip()
                    if not name or name == 'None': continue
                    
                    match = df_seeds[df_seeds['이름_key'] == name]
                    if not match.empty:
                        s_row = match.iloc[0]
                        if t_dob:
                            cell = target_sheet.cell(row=r, column=col_indices[t_dob])
                            if not cell.value or str(cell.value).strip() == '':
                                cell.value = s_row['생년월일_val']
                                fill_count += 1
                        if t_comp:
                            cell = target_sheet.cell(row=r, column=col_indices[t_comp])
                            if not cell.value or str(cell.value).strip() == '':
                                cell.value = s_row['업체_val']
                        if t_job:
                            cell = target_sheet.cell(row=r, column=col_indices[t_job])
                            if not cell.value or str(cell.value).strip() == '':
                                cell.value = s_row['직종_val']

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                today = datetime.now().strftime("%Y%m%d")
                st.success(f"✅ 병합 완료! (총 {fill_count}개의 항목을 확인/업데이트 했습니다.)")
                st.download_button(
                    label="📥 병합된 결과 파일 다운로드",
                    data=output,
                    file_name=f"{today}_병합결과_{target_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"처리 중 예상치 못한 오류가 발생했습니다: {e}")
